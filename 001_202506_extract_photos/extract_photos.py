import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

def extract_images_from_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        image_loader = SheetImageLoader(ws)

        # 获取Excel所在文件夹路径
        output_dir = os.path.dirname(excel_path)

        row = 4
        extracted = 0
        while True:
            name_cell = f"C{row}"
            photo_cell = f"D{row}"

            name = ws[name_cell].value
            if not name:
                break  # 数据读取结束

            if image_loader.image_in(photo_cell):
                image = image_loader.get(photo_cell)
                filename = f"{name}_大头照.jpg"
                save_path = os.path.join(output_dir, filename)
                image.save(save_path)
                print(f"已保存: {save_path}")
                extracted += 1
            else:
                print(f"未找到{name}的照片")
            row += 1

        messagebox.showinfo("完成", f"已提取 {extracted} 张照片，保存在\n{output_dir}")
    except Exception as e:
        messagebox.showerror("出错", f"提取失败：{str(e)}")

def main():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="请选择包含照片的Excel文件",
        filetypes=[("Excel 文件", "*.xlsx *.xlsm")]
    )
    if not file_path:
        sys.exit()  # 用户取消

    extract_images_from_excel(file_path)

if __name__ == "__main__":
    main()
